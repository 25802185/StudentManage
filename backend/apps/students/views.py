import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework import status as http_status
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl

from .models import Student, InfoChangeRequest
from .serializers import (
    StudentSerializer, StudentCreateSerializer, StudentUpdateSerializer,
    InfoChangeRequestSerializer,
)
from apps.users.permissions import IsAdminOrTeacher, IsAdmin
from apps.users.models import User


class StudentViewSet(ModelViewSet):
    queryset = Student.objects.select_related('user', 'class_ref').all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['class_ref', 'gender', 'status']
    search_fields = ['name', 'student_no']
    ordering_fields = ['student_no', 'name']

    def get_permissions(self):
        return [IsAdminOrTeacher()]

    def get_serializer_class(self):
        if self.action == 'create':
            return StudentCreateSerializer
        if self.action in ['update', 'partial_update']:
            return StudentUpdateSerializer
        return StudentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == 'teacher' and hasattr(user, 'teacher_profile'):
            qs = qs.filter(class_ref=user.teacher_profile.class_ref)
        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        students = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '学生信息'
        headers = ['学号', '姓名', '性别', '年龄', '班级', '电话', '邮箱', '地址']
        ws.append(headers)
        for s in students:
            ws.append([
                s.student_no, s.name, s.get_gender_display(), s.age,
                s.class_ref.name if s.class_ref else '', s.phone, s.email, s.address,
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=students_{datetime.now():%Y%m%d}.xlsx'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传文件'}, status=http_status.HTTP_400_BAD_REQUEST)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        for row in rows:
            student_no, name, gender_str, age, class_name, phone, email, address = (
                str(row[0]), str(row[1]), row[2], row[3], row[4],
                row[5] or '', row[6] or '', row[7] or '',
            )
            if Student.objects.filter(student_no=student_no).exists():
                continue
            gender = 'male' if gender_str == '男' else 'female'
            from apps.classes.models import ClassInfo
            class_ref = ClassInfo.objects.filter(name=class_name).first()
            username = student_no
            password = username[-6:]
            user = User.objects.create_user(
                username=username, password=password, role='student', is_active=True,
            )
            Student.objects.create(
                user=user, student_no=student_no, name=name,
                gender=gender, age=int(age) if age else None,
                class_ref=class_ref, phone=str(phone), email=str(email), address=str(address),
            )
            created += 1
        return Response({'detail': f'成功导入 {created} 条'})


class InfoChangeRequestViewSet(ModelViewSet):
    queryset = InfoChangeRequest.objects.select_related('student', 'reviewer').all()
    serializer_class = InfoChangeRequestSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == 'student':
            return qs.filter(student__user=user)
        if user.role == 'teacher' and hasattr(user, 'teacher_profile'):
            return qs.filter(student__class_ref=user.teacher_profile.class_ref)
        return qs

    def perform_create(self, serializer):
        student = Student.objects.get(user=self.request.user)
        student.status = 'pending'
        student.save()
        serializer.save(student=student)

    @action(detail=True, methods=['put'], url_path='approve')
    def approve(self, request, pk=None):
        change = self.get_object()
        change.status = 'approved'
        change.reviewer = request.user
        change.review_time = timezone.now()
        change.save()
        student = change.student
        setattr(student, change.field_name, change.new_value)
        student.status = 'normal'
        student.save()
        return Response({'detail': '审核通过'})

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        change = self.get_object()
        change.status = 'rejected'
        change.reviewer = request.user
        change.review_time = timezone.now()
        change.remark = request.data.get('remark', '')
        change.save()
        student = change.student
        student.status = 'normal'
        student.save()
        return Response({'detail': '已驳回'})
